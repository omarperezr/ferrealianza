import fitz  # PyMuPDF
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image as PILImage
import io

def extract_full_catalog(pdf_path, excel_path):
    print(f"Abriendo documento: {pdf_path}")
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        print(f"Error al abrir el PDF: {e}")
        return

    # 1. Crear el libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catálogo LDP"

    # 2. Configurar encabezados
    # Asumimos hasta 5 columnas de texto basadas en el PDF + 1 de Foto
    headers = ["Código/Categoría", "Descripción del Producto", "Empaque / Detalles", "Precio / Ref 1", "Precio / Ref 2", "Foto"]
    ws.append(headers)
    
    # Estilos para encabezados
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="4F81BD", fill_type="solid")

    ws.column_dimensions['F'].width = 15 # Ancho de columna para la foto
    row_excel = 2
    
    # 3. Iterar por cada página del documento
    for page_num in range(len(doc)):
        print(f"Procesando página {page_num + 1} de {len(doc)}...")
        page = doc.load_page(page_num)
        
        # A. Extraer todos los bloques de texto con sus coordenadas
        blocks = page.get_text("blocks")
        text_elements = []
        for b in blocks:
            if b[6] == 0:  # b[6] == 0 significa que es un bloque de texto
                text = b[4].strip()
                if text: # Ignorar bloques completamente vacíos
                    text_elements.append({
                        "type": "text",
                        "bbox": b[:4], # (x0, y0, x1, y1)
                        "y_center": (b[1] + b[3]) / 2, # Centro vertical del bloque
                        "content": text
                    })
        
        # B. Extraer imágenes con sus coordenadas en la página
        image_elements = []
        try:
            # get_image_info devuelve las ubicaciones reales (bbox) de las imágenes en la página
            images_info = page.get_image_info(xrefs=True)
            for img in images_info:
                bbox = img["bbox"]
                image_elements.append({
                    "type": "image",
                    "bbox": bbox,
                    "y_center": (bbox[1] + bbox[3]) / 2,
                    "xref": img["xref"]
                })
        except Exception as e:
            print(f"Advertencia: No se pudieron extraer posiciones de imágenes en pág {page_num + 1}. {e}")
        
        # C. Unir textos e imágenes y agruparlos por FILAS (basado en su posición Y)
        all_elements = text_elements + image_elements
        if not all_elements:
            continue
            
        # Ordenar todos los elementos de arriba hacia abajo
        all_elements.sort(key=lambda e: e["y_center"])
        
        rows = []
        current_row = []
        current_y = all_elements[0]["y_center"]
        
        # TOLERANCIA: Píxeles de diferencia vertical permitidos para considerar que están en la misma fila
        TOLERANCE = 25 
        
        for el in all_elements:
            if abs(el["y_center"] - current_y) <= TOLERANCE:
                current_row.append(el)
            else:
                rows.append(current_row)
                current_row = [el]
                current_y = el["y_center"]
        if current_row:
            rows.append(current_row)
            
        # D. Procesar cada fila detectada y enviarla a Excel
        for row in rows:
            # Ordenar los elementos de la fila de izquierda a derecha (coordenada X)
            row.sort(key=lambda e: e["bbox"][0])
            
            # Filtro: Omitir filas que sean los encabezados repetidos de las páginas del PDF
            row_texts = [e["content"] for e in row if e["type"] == "text"]
            text_joined = " ".join(row_texts).upper()
            if "LISTA DE MAYO" in text_joined or "CAT CODIGO" in text_joined or ("PRODUCTO" in text_joined and "EMP" in text_joined):
                continue
                
            col_idx = 1
            img_added = False
            
            # Escribir elementos en las columnas correspondientes
            for el in row:
                if el["type"] == "text":
                    if col_idx <= 5: # Límite de 5 columnas de texto antes de la foto
                        # Limpiar saltos de línea extraños del PDF
                        clean_text = el["content"].replace('\n', ' ').strip()
                        ws.cell(row=row_excel, column=col_idx, value=clean_text)
                        col_idx += 1
                
                elif el["type"] == "image":
                    # Procesar e insertar la imagen en la columna F
                    xref = el["xref"]
                    try:
                        base_image = doc.extract_image(xref)
                        image_bytes = base_image["image"]
                        
                        img_stream = io.BytesIO(image_bytes)
                        pil_img = PILImage.open(img_stream)
                        
                        # Convertir a formato RGB para evitar errores con imágenes CMYK o paletas raras al guardar en PNG
                        if pil_img.mode in ("RGBA", "P", "CMYK"):
                            pil_img = pil_img.convert("RGB")
                            
                        pil_img.thumbnail((80, 80)) # Redimensionar foto para Excel
                        
                        out_stream = io.BytesIO()
                        pil_img.save(out_stream, format="PNG")
                        out_stream.seek(0)
                        
                        xl_img = ExcelImage(out_stream)
                        
                        # Ajustar alto de celda e insertar
                        ws.row_dimensions[row_excel].height = 65
                        ws.add_image(xl_img, f"F{row_excel}")
                        img_added = True
                    except Exception as e:
                        print(f"Error procesando imagen xref {xref}: {e}")
            
            # Solo avanzamos a la siguiente fila de Excel si realmente se extrajo texto o una imagen
            if row_texts or img_added:
                row_excel += 1

    # 4. Guardar archivo final
    wb.save(excel_path)
    print(f"¡Extracción exitosa! Archivo guardado en: {excel_path}")

if __name__ == "__main__":
    pdf_file = "lpd.pdf" # Reemplaza con tu archivo PDF
    excel_file = "Catalogo_Extraido_Completo.xlsx"
    extract_full_catalog(pdf_file, excel_file)